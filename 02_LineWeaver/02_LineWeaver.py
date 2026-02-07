import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
import base64
from tempfile import NamedTemporaryFile
import os
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO

# 言語切替
lang = st.selectbox("Language / 言語を選択してください", ["日本語", "English"])
texts = {
    "日本語": {
        "title1": "Skyrim Mod 翻訳支援ツール",
        "title2": "02 LineWeaver",
        "step1_title1": "ステップ1：",
        "step1_title2": "SSEEdit からマージ用の csv（02LW_XXX.csv）を出力してください",
        "step1_desc": "スクリプトが必要なら、↓からダウンロード",
        "pas_download": "📥 02LW_step1.pas をダウンロード",
        "step2_title1": "ステップ2：",
        "step2_title2": "02LW_XXX.csv と xml をアップロードしてください",
        "csv_upload": "▼02LW_XXX.csv と xml をここにドロップ",
        "convert_btn": "xlsx に変換開始",
        "xlsx_download": "📥 XLSXをダウンロード",
        "step3_title1": "ステップ3：",
        "step3_title2": "翻訳が終わった xlsx をアップロードしてください",
        "xlsx_upload": "▼翻訳が終わった xlsx をここにドロップ",
        "convert_xml_btn": "xml に変換開始",
        "xml_download": "📥 XMLをダウンロード",
        "tutorial": "使い方はこちら。よく読んでからご利用くださいね"
    },
    "English": {
        "title1": "Skyrim MOD Translation Assistant",
        "title2": "02 LineWeaver",
        "step1_title1": "Step 1:",
        "step1_title2": "Export merge csv (02LW_XXX.csv) from SSEEdit",
        "step1_desc": "If script is needed, download below",
        "pas_download": "📥 Download 02LW_step1.pas",
        "step2_title1": "Step 2:",
        "step2_title2": "Upload 02LW_XXX.csv and xml",
        "csv_upload": "▼Drop your 02LW_XXX.csv and xml here",
        "convert_btn": "Start conversion to xlsx",
        "xlsx_download": "📥 Download XLSX",
        "step3_title1": "Step 3:",
        "step3_title2": "Upload translated xlsx",
        "xlsx_upload": "▼Drop your translated xlsx here",
        "convert_xml_btn": "Start conversion to XML",
        "xml_download": "📥 Download XML",
        "tutorial": "Click here for tutorial (read before use)"
    }
}
t = texts[lang]

# CSS
st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: 'Meiryo', sans-serif;
}
hr {
    border: none;
    border-top: 2px solid #888;
    margin: 40px 0;
}
h1, h2, h3 {
    margin-bottom: 10px;
}
.section-title {
    font-size: 24px;
    font-weight: bold;
}
.sub-title {
    font-size: 18px;
    margin-bottom: 10px;
}
.stDownloadButton > button {
    background-color: #45818e !important;
    color: white;
    font-weight: bold;
    border-radius: 10px;
    padding: 10px 40px;
}
.stButton > button {
    background-color: #3d85c6 !important;
    color: white;
    font-weight: bold;
    border-radius: 10px;
    padding: 10px 40px;
}
.dropbox {
    background-color: transparent !important;
    border: none !important;
    padding: 0 !important;
    margin: 10px 0;
    color: inherit;
}
</style>
""", unsafe_allow_html=True)

# タイトル
st.markdown(f"<h1 style='text-align:center;'>{t['title1']}<br>{t['title2']}</h1>", unsafe_allow_html=True)
st.markdown(f"<div style='text-align:center; margin: 20px 0;'><a href='https://note.com/ofutonomy/n/nbe796dd6ffe5' target='_blank' style='background-color:#0b5394; color:white; padding:12px 24px; border-radius:8px; text-decoration:none; font-weight:bold; display:inline-block;'>{t['tutorial']}</a></div>", unsafe_allow_html=True)

# ステップ1
st.markdown("<hr>", unsafe_allow_html=True)
st.markdown(f"<div class='section-title'>{t['step1_title1']}<br><span class='sub-title'>{t['step1_title2']}</span></div>", unsafe_allow_html=True)
st.markdown(f"<p>{t['step1_desc']}</p>", unsafe_allow_html=True)

pas_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "02LW_step1.pas")
if os.path.exists(pas_path):
    with open(pas_path, "rb") as f:
        st.download_button(label=t["pas_download"], data=f, file_name="02LW_step1.pas", mime="text/plain")
else:
    st.warning("02LW_step1.pas が見つかりません（ローカル実行をおすすめします）")

# ステップ2：02LW_XXX.csv と xml をアップロードしてマージ出力
st.markdown("<hr>", unsafe_allow_html=True)
st.markdown(f"<div class='section-title'>{t['step2_title1']}<br><span class='sub-title'>{t['step2_title2']}</span></div>", unsafe_allow_html=True)
st.markdown(f"<div class='dropbox'>{t['csv_upload']}</div>", unsafe_allow_html=True)
csv_file = st.file_uploader("CSV", type=["csv"], key="csv_input")
xml_file = st.file_uploader("XML", type=["xml"], key="xml_input")

def detect_formid_or_none(val):
    if isinstance(val, str) and len(val) == 8:
        try:
            int(val, 16)
            return val.upper()
        except ValueError:
            return np.nan
    return np.nan

def resolve_formid(row):
    if pd.notna(row["FormID_direct"]) and str(row["FormID_direct"]).strip() != "":
        return row["FormID_direct"]
    elif pd.notna(row["FormID_x"]) and str(row["FormID_x"]).strip() != "":
        return row["FormID_x"]
    return ""

if st.button(t["convert_btn"], key="to_xlsx") and csv_file and xml_file:
    csv1 = pd.read_csv(csv_file, dtype=str).fillna("")
    xml_content = xml_file.read().decode("utf-8")
    root = ET.fromstring(xml_content)

    rows = []
    for string in root.find("Content").findall("String"):
        rec_elem = string.find("REC")
        row = {
            "List": string.attrib.get("List", ""),
            "Partial": string.attrib.get("Partial", ""),
            "EDID": string.findtext("EDID", ""),
            "REC": string.findtext("REC", ""),
            "REC_id": rec_elem.attrib.get("id", "") if rec_elem is not None else "",
            "REC_idMax": rec_elem.attrib.get("idMax", "") if rec_elem is not None else "",
            "Source": string.findtext("Source", ""),
            "Dest": string.findtext("Dest", "")
        }
        rows.append(row)
    csv2 = pd.DataFrame(rows)

    csv2["EDID_clean"] = csv2["EDID"].astype(str).str.replace(r"[\[\]]", "", regex=True)
    csv2["FormID_direct"] = csv2["EDID_clean"].apply(detect_formid_or_none)

    csv3 = pd.merge(
        csv2,
        csv1[["FormID", "EDID"]],
        left_on="EDID_clean",
        right_on="EDID",
        how="left",
        suffixes=("", "_from_edid")
    ).rename(columns={"FormID": "FormID_x", "EDID_from_edid": "EDID_from_edid"})

    csv3["FormID_final"] = csv3.apply(resolve_formid, axis=1)
    csv3["FormID_final"] = csv3["FormID_final"].apply(lambda x: f"{int(x, 16):08X}" if str(x).strip() != "" else "")

    csv1["FormID"] = csv1["FormID"].astype(str).str.zfill(8).str.upper()
    csv3["FormID_final"] = csv3["FormID_final"].astype(str).str.zfill(8).str.upper()
    # Merge Speaker and new columns (ESP_Order, ParentDIAL, ParentDIAL_EDID) if available
    merge_cols = ["FormID", "Speaker"]
    for col in ["ESP_Order", "ParentDIAL", "ParentDIAL_EDID"]:
        if col in csv1.columns:
            merge_cols.append(col)
    csv4 = pd.merge(
        csv3,
        csv1[merge_cols],
        left_on="FormID_final",
        right_on="FormID",
        how="left"
    ).rename(columns={"FormID": "FormID_y"})

    csv4["REC_id_sort"] = pd.to_numeric(csv4["REC_id"], errors="coerce").fillna(-1)
    if "ESP_Order" in csv4.columns:
        # ESP structure order: preserves DIAL→INFO grouping
        csv4["ESP_Order_sort"] = pd.to_numeric(csv4["ESP_Order"], errors="coerce").fillna(999999)
        csv4 = csv4.sort_values(by=["ESP_Order_sort", "REC_id_sort"], na_position="last").drop(columns=["ESP_Order_sort", "REC_id_sort"])
    else:
        # Fallback: FormID order (old .pas without ESP_Order)
        csv4["FormID_final_sort"] = csv4["FormID_final"].fillna("")
        csv4 = csv4.sort_values(by=["FormID_final_sort", "REC_id_sort"], na_position="last").drop(columns=["FormID_final_sort", "REC_id_sort"])

    # Drop intermediate merge columns, keep only useful ones
    output_cols = ["List", "Partial", "EDID", "REC", "REC_id", "REC_idMax",
                   "FormID_final", "Speaker", "Source", "Dest"]
    # Add new columns if available (from updated .pas)
    for col in ["ESP_Order", "ParentDIAL", "ParentDIAL_EDID"]:
        if col in csv4.columns:
            output_cols.append(col)
    csv4 = csv4[[c for c in output_cols if c in csv4.columns]]

    plugin_base = csv_file.name.replace("02LW_", "").replace(".csv", "")
    output_filename = f"02LW_{plugin_base}.xlsx"

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        csv4.to_excel(tmp.name, index=False)

        wb = load_workbook(tmp.name)
        ws = wb.active
        fill = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")

        def is_english(text):
            return isinstance(text, str) and all(ord(c) < 128 for c in text)

        excluded_rec = {
            "HDPT:FULL", "FURN:FULL", "CLAS:FULL", "DOOR:FULL", "FACT:FULL", "CLFM:FULL",
            "TREE:FULL", "MGEF:FULL", "RACE:FULL", "EXPL:FULL", "HAZD:FULL", "BPTD:BPTN"
        }
        mask = (~csv4["REC"].isin(excluded_rec)) & (csv4["Source"] == csv4["Dest"])

        # Find the Dest column position dynamically (not hardcoded)
        dest_col_idx = list(csv4.columns).index("Dest") + 1  # 1-indexed for openpyxl
        dest_col_letter = get_column_letter(dest_col_idx)
        for i, match in enumerate(mask, start=2):
            if match:
                ws[f"{dest_col_letter}{i}"].fill = fill
        wb.save(tmp.name)

        with open(tmp.name, "rb") as f:
            st.download_button(t["xlsx_download"], f, file_name=output_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ステップ3
st.markdown("<hr>", unsafe_allow_html=True)
st.markdown(f"<div class='section-title'>{t['step3_title1']}<br><span class='sub-title'>{t['step3_title2']}</span></div>", unsafe_allow_html=True)
st.markdown(f"<div class='dropbox'>{t['xlsx_upload']}</div>", unsafe_allow_html=True)
xlsx_file = st.file_uploader("XLSX", type=["xlsx"], key="xlsx_input")

# 翻訳先言語を選択
dest_lang = st.selectbox(
    "Select destination language (for xTranslator)",
    ["japanese", "german", "french", "spanish", "russian", "polish", "italian", "chinese", "korean", "portuguese", "turkish"],
    index=0
)

if st.button(t["convert_xml_btn"], key="to_xml") and xlsx_file:
    df = pd.read_excel(xlsx_file, dtype=str).fillna("")
    plugin_base = xlsx_file.name.replace("02LW_", "").replace(".xlsx", "")

    # XML構築
    root = ET.Element("SSTXMLRessources")
    params = ET.SubElement(root, "Params")
    ET.SubElement(params, "Addon").text = f"{plugin_base}.esp"
    ET.SubElement(params, "Source").text = "english"
    ET.SubElement(params, "Dest").text = dest_lang  # ← ここで選択された言語を使用
    ET.SubElement(params, "Version").text = "2"

    content = ET.SubElement(root, "Content")

    for _, row in df.iterrows():
        s_elem = ET.SubElement(content, "String", List=row.get("List", ""), Partial=row.get("Partial", ""))
        ET.SubElement(s_elem, "EDID").text = row.get("EDID", "")
        rec_elem = ET.SubElement(s_elem, "REC", id=row.get("REC_id", ""), idMax=row.get("REC_idMax", ""))
        rec_elem.text = row.get("REC", "")
        ET.SubElement(s_elem, "Source").text = row.get("Source", "")
        ET.SubElement(s_elem, "Dest").text = row.get("Dest", "")

    # XML書き出し
    xml_bytesio = BytesIO()
    tree = ET.ElementTree(root)
    tree.write(xml_bytesio, encoding="utf-8", xml_declaration=True)
    xml_bytes = xml_bytesio.getvalue()

    # ダウンロードボタン
    st.download_button(
        label=t["xml_download"],
        data=xml_bytes,
        file_name=f"{plugin_base}.xml",
        mime="application/xml"
    )
